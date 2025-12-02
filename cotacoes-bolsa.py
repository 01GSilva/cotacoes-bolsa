import yfinance as yf
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import os
import tkinter as tk
from tkinter import ttk

arquivo_excel = "cotacoes_bolsa.xlsx"

# ---------------- AÇOES NACIONAIS E FIIS-------------------


def solicitar_quantidades(lista):
    quantities = {}

    def confirmar():
        for ticker, entry in entries.items():
            try:
                q = float(entry.get())
            except:
                q = 0
            quantities[ticker] = q
        root.destroy()

    root = tk.Tk()
    root.title("Quantidades de Ativos")

    tk.Label(root, text="Digite as quantidades de cada ativo:", font=(
        "Arial", 12, "bold")).grid(row=0, column=0, columnspan=2, pady=10)

    entries = {}
    row_index = 1

    for ticker in lista:
        tk.Label(root, text=ticker, font=("Arial", 10)).grid(
            row=row_index, column=0, padx=10, pady=5, sticky="w")

        entry = ttk.Entry(root, width=10)
        entry.grid(row=row_index, column=1, padx=10)
        entry.insert(0, "0")

        entries[ticker] = entry
        row_index += 1

    btn = ttk.Button(root, text="Confirmar", command=confirmar)
    btn.grid(row=row_index, column=0, columnspan=2, pady=15)

    root.mainloop()

    return quantities


def coletar_precos(lista, nome_coluna):

    quantidades = solicitar_quantidades(lista)

    dados = {}

    for item in lista:
        tk_item = yf.Ticker(item)
        hist = tk_item.history(period="1d")

        if hist.empty:
            print(f"Atenção: {item} não retornou dados.")
            preco = None
        else:
            preco = hist["Close"].iloc[-1]

        quantidade = quantidades.get(item, 0)
        total_investido = preco * quantidade if preco is not None else None

        dados[item] = {
            "Preço Atual": preco,
            "Quantidade": quantidade,
            "Total Investido": total_investido,
            "Atualizado em": datetime.now().strftime("%d/%m/%Y %H:%M:%S")}

    df = pd.DataFrame({
        nome_coluna: list(dados.keys()),
        "Preço Atual": [d["Preço Atual"] for d in dados.values()],
        "Quantidade": [d["Quantidade"] for d in dados.values()],
        "Total Investido": [d["Total Investido"] for d in dados.values()],
        "Atualizado em": [d["Atualizado em"] for d in dados.values()]
    })

    # Garantir tipos numéricos
    df["Preço Atual"] = pd.to_numeric(
        df["Preço Atual"], errors="coerce").fillna(0.0)
    df["Quantidade"] = pd.to_numeric(
        df["Quantidade"], errors="coerce").fillna(0.0)
    df["Total Investido"] = pd.to_numeric(
        df["Total Investido"], errors="coerce").fillna(0.0)

    return df

# ---------------- AÇÕES INTERNACIONAIS -------------------


def pedir_valores_internacionais():
    def confirmar():
        try:
            hist = yf.Ticker('USDBRL=X').history(period='1d')
            if hist.empty:
                raise RuntimeError(
                    'Não foi possivel coletar a cotação do dolar')
            cotacao = float(hist['Close'].iloc[0])

            def ler(entry):
                txt = entry.get().strip()
                return float(txt.replace('.', ',')) if txt else 0.0

            valores = {
                'SPHD': ler(entry_sphd),
                'SDY': ler(entry_sdy),
                'PFF': ler(entry_pff),
                'SDIV': ler(entry_sdiv)
            }

            agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

            linhas = []
            for ativo, usd in valores.items():
                linhas.append({
                    'Ação:': ativo,
                    'Valor em USD': usd,
                    'Cotação Dólar': cotacao,
                    'Valor em Reais': usd * cotacao,
                    'Atualizado em': agora
                })

            df_internacionais = pd.DataFrame(linhas)

            resultado.append(df_internacionais)
            root.destroy()

        except Exception as e:
            print(f'Erro ao confirmar valores\n{e}')

    root = tk.Tk()
    root.title('Investimentos Internacionais (USD)')

    ttk.Label(root, text='Digite os valores investidos em USD').grid(
        row=0, column=0, columnspan=2, pady=10)

    ttk.Label(root, text='SPHD:').grid(row=1, column=0, sticky='e')
    entry_sphd = ttk.Entry(root)
    entry_sphd.grid(row=1, column=1)
    ttk.Label(root, text='SDY:').grid(row=2, column=0, sticky='e')
    entry_sdy = ttk.Entry(root)
    entry_sdy.grid(row=2, column=1)
    ttk.Label(root, text='PFF:').grid(row=3, column=0, sticky='e')
    entry_pff = ttk.Entry(root)
    entry_pff.grid(row=3, column=1)
    ttk.Label(root, text='SDIV:').grid(row=4, column=0, sticky='e')
    entry_sdiv = ttk.Entry(root)
    entry_sdiv.grid(row=4, column=1)

    ttk.Button(root, text='Confirmar', command=confirmar).grid(
        row=5, column=0, columnspan=2, pady=10)

    resultado = []
    root.mainloop()

    return resultado[0] if resultado else None

# ---------------- RENDA FIXA -------------------


def pedir_valores_renda_fixa():
    def confirmar():
        try:
            txt = entry.get().strip()
            valor = float(txt.replace('.', ',')) if txt else 0.0

            agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

            linha = {
                'Ação': 'Renda Fixa',
                'Valor em USD': None,
                'Cotação Dólar': None,
                'Valor em Reais': valor,
                'Atualizado em': agora
            }

            df = pd.DataFrame([linha])

            resultado.append(df)
            root.destroy()

        except Exception as e:
            print('Erro', f'Erro ao confirmar renda fixa:\n{e}')

    root = tk.Tk()
    root.title('Valor investido em Renda Fixa (R$)')

    ttk.Label(root, text='Digite o valor total investido em Renda Fixa (R$):').grid(
        row=0, column=0, padx=10, pady=10)

    entry = ttk.Entry(root)
    entry.grid(row=1, column=0, padx=10)

    ttk.Button(root, text='Confirmar', command=confirmar).grid(
        row=2, column=0, pady=10)

    resultado = []
    root.mainloop()

    return resultado[0] if resultado else None

# ---------------- LISTA DE ATIVOS -------------------


acoes_na = ['ITSA4.SA', 'BBAS3.SA', 'ABEV3.SA', 'VIVT3.SA',
            'VALE3.SA', 'BRAP4.SA', 'CMIG4.SA', 'EGIE3.SA',
            'CSMG3.SA', 'SAPR4.SA']

fiis = ['FLMA11.SA', 'TGAR11.SA', 'GGRC11.SA', 'PCIP11.SA',
        'VISC11.SA', 'BTAL11.SA', 'PVBI11.SA', 'HSML11.SA',
        'RFOF11.SA', 'BTHF11.SA', 'VINO11.SA', 'MXRF11.SA']

# ---------------- COLETA DOS DADOS -------------------

df_na = coletar_precos(acoes_na, "Ação")
df_int = pedir_valores_internacionais()
df_fii = coletar_precos(fiis, "FII")
df_rf = pedir_valores_renda_fixa()

# ---------------- PORCENTAGENS -------------------

total_na = df_na['Total Investido'].sum()
total_int = df_int['Valor em Reais'].sum()
total_fii = df_fii['Total Investido'].sum()
total_rf = df_rf['Valor em Reais'].sum()

total_geral = total_na + total_int + total_fii + total_rf

if total_geral == 0:
    print('Nenhum valor total encontrado. Impossivel calcular porcentagem')
else:
    pct_na = (total_na/total_geral)*100
    pct_int = (total_int/total_geral)*100
    pct_fii = (total_fii/total_geral)*100
    pct_rf = (total_rf/total_geral)*100

porcentagens = {
    'Nacionais': [pct_na],
    'Internacionais': [pct_int],
    'Fiis': [pct_fii],
    'Renda Fixa': [pct_rf]
}

df_pct = pd.DataFrame(porcentagens)

# ---------------- EXPORTAÇÃO -------------------

with pd.ExcelWriter(arquivo_excel, engine="openpyxl", mode="w") as writer:
    df_na.to_excel(writer, sheet_name="Ações Nacionais", index=False)
    df_int.to_excel(writer, sheet_name="Ações Internacionais", index=False)
    df_fii.to_excel(writer, sheet_name="FIIs", index=False)
    df_rf.to_excel(writer, sheet_name="Renda Fixa", index=False)
    df_pct.to_excel(writer, sheet_name="Porcentagens", index=False)

# ---------------- FORMATAÇÃO -------------------

wb = load_workbook(arquivo_excel)

for aba in ["Ações Nacionais", "Ações Internacionais", "FIIs", "Renda Fixa"]:
    ws = wb[aba]

    try:
        for cell in ws['B'][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = 'R$ #,##0.00'
    except Exception:
        pass

    try:
        for cell in ws['D'][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = 'R$ #,##0.00'
    except Exception:
        pass

wb.save(arquivo_excel)

print("Arquivo Excel gerado com sucesso!")
print("Caminho completo:", os.path.abspath(arquivo_excel))
